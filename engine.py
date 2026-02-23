import os
import re
import unicodedata
import hashlib
import datetime
from difflib import SequenceMatcher
from typing import List, Tuple, Optional

import numpy as np
import pandas as pd
import openpyxl


# ============================================================
# Utilidades de normalización
# ============================================================

def _norm_txt(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    s = str(x)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.upper().strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^A-Z0-9 ]", "", s)
    return s


def normalize_dni(val):
    """Extrae un DNI de 8 dígitos desde cualquier formato común (float/string)."""
    if pd.isna(val):
        return np.nan
    s = str(val).strip()
    s = re.sub(r"\.0$", "", s)
    digs8 = re.findall(r"\d{8}", s)
    if digs8:
        return digs8[0]
    digs = re.findall(r"\d+", s)
    if digs:
        d = digs[0]
        return d[:8] if len(d) >= 8 else np.nan
    return np.nan


def _robust_datetime_from_any(series: pd.Series) -> pd.Series:
    """
    Parse dates robustly from:
    - Excel serial numbers (days since 1899-12-30)
    - strings (dd/mm/yyyy, yyyy-mm-dd, etc.)
    - datetime objects
    Returns datetime64[ns] with NaT for invalid.
    """
    s = series.copy()

    # Numeric -> try Excel serial conversion if looks like a serial date
    num = pd.to_numeric(s, errors="coerce")
    mask_serial = num.notna() & (num.between(20000, 60000))  # ~1954..2064
    dt = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    if mask_serial.any():
        dt.loc[mask_serial] = pd.to_datetime(num.loc[mask_serial], unit="D", origin="1899-12-30", errors="coerce")

    # Non-serial -> normal parse (dayfirst)
    mask_other = ~mask_serial
    if mask_other.any():
        dt.loc[mask_other] = pd.to_datetime(s.loc[mask_other], errors="coerce", dayfirst=True)

    # Remove typical artifacts
    dt = dt.mask(dt.dt.year == 1970)

    return dt


def _parse_dates_two_modes(series: pd.Series) -> tuple[pd.Series, pd.Series, pd.Series]:
    """
    Devuelve (dt_dayfirst, dt_monthfirst, time_only_mask).
    - Números: se interpretan como serial de Excel si caen en rango razonable.
    - Strings: se parsean en ambos modos.
    """
    s = series.copy()

    # time-only string -> NaT
    time_only = s.astype(str).str.match(r"^\s*\d{1,2}:\d{2}(:\d{2})?\s*$", na=False)

    num = pd.to_numeric(s, errors="coerce")
    mask_serial = num.notna() & (num.between(20000, 60000))  # ~1954..2064

    dt1 = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")
    dt2 = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    if mask_serial.any():
        dt_serial = pd.to_datetime(num.loc[mask_serial], unit="D", origin="1899-12-30", errors="coerce")
        dt1.loc[mask_serial] = dt_serial
        dt2.loc[mask_serial] = dt_serial

    mask_other = ~mask_serial
    if mask_other.any():
        dt1.loc[mask_other] = pd.to_datetime(s.loc[mask_other], errors="coerce", dayfirst=True)
        dt2.loc[mask_other] = pd.to_datetime(s.loc[mask_other], errors="coerce", dayfirst=False)

    # clean artifacts
    dt1 = dt1.mask(time_only).mask(dt1.dt.year == 1970)
    dt2 = dt2.mask(time_only).mask(dt2.dt.year == 1970)

    return dt1, dt2, time_only


def standardize_date_contextual(series: pd.Series, report: list[str], context_label: str) -> pd.Series:
    """
    Normaliza fechas a dd/mm/yyyy, eligiendo entre dayfirst=True o False para valores ambiguos
    basado en continuidad (mes dominante / mediana).
    No hardcodea año ni mes; aprende del propio archivo/hoja.

    Registra en reporte cuántos swaps hizo.
    """
    dt1, dt2, _ = _parse_dates_two_modes(series)

    # Ambiguo: ambos válidos y diferentes
    amb = dt1.notna() & dt2.notna() & (dt1 != dt2)

    # Fechas "no ambiguas": dt2 NaT o iguales
    stable = dt1.notna() & (~amb)

    # Contexto: mes dominante y mediana
    ctx_src = dt1[stable]
    if len(ctx_src) < 25:
        ctx_src = dt1[dt1.notna()]

    if len(ctx_src) == 0:
        chosen = dt1
        return chosen.dt.strftime("%d/%m/%Y").where(chosen.notna(), np.nan)

    # mes dominante
    ym = ctx_src.dt.to_period("M").astype(str)
    dom = ym.value_counts().idxmax()  # 'YYYY-MM'
    dom_year, dom_month = map(int, dom.split("-"))

    median = ctx_src.median()
    # ventana de continuidad: +/- 60 días alrededor de la mediana
    lo = median - pd.Timedelta(days=60)
    hi = median + pd.Timedelta(days=60)

    chosen = dt1.copy()
    swaps = 0
    kept_amb = 0

    if amb.any():
        d1 = dt1[amb]
        d2 = dt2[amb]

        # condiciones de "tiene sentido"
        in_dom_1 = (d1.dt.year == dom_year) & (d1.dt.month == dom_month)
        in_dom_2 = (d2.dt.year == dom_year) & (d2.dt.month == dom_month)

        in_win_1 = (d1 >= lo) & (d1 <= hi)
        in_win_2 = (d2 >= lo) & (d2 <= hi)

        # regla:
        # 1) si solo uno cae en mes dominante -> elegir ese
        pick2 = (in_dom_2 & ~in_dom_1)
        pick1 = (in_dom_1 & ~in_dom_2)

        # 2) si ninguno/ambos en dom, usar ventana de continuidad
        pick2 |= (~pick1 & ~pick2 & in_win_2 & ~in_win_1)

        # 3) si ambos en ventana o ambos fuera, elegir el más cercano a la mediana
        tie = ~pick1 & ~pick2
        if tie.any():
            dist1 = (d1[tie] - median).abs()
            dist2 = (d2[tie] - median).abs()
            pick2_tie = dist2 < dist1
            idx2 = pick2_tie[pick2_tie].index
            pick2.loc[idx2] = True

        idx2 = pick2[pick2].index
        chosen.loc[idx2] = dt2.loc[idx2]
        swaps = len(idx2)
        kept_amb = int(amb.sum() - swaps)

    if swaps:
        report.append(f"[DATE] {context_label} | swaps_ambiguous={swaps} kept_ambiguous={kept_amb} | dom={dom_year:04d}-{dom_month:02d} median={median.date()}")
    return chosen.dt.strftime("%d/%m/%Y").where(chosen.notna(), np.nan)


def standardize_date(series: pd.Series) -> pd.Series:
    """
    Normaliza a dd/mm/yyyy.
    - Soporta datetime, seriales de Excel y strings.
    - Si el valor era solo hora (ej. '08:30'), lo deja en NaN.
    - Si cae en 01/01/1970 (artefacto típico), lo deja en NaN.
    """
    s = series.copy()

    # time-only string -> NaN
    time_only = s.astype(str).str.match(r"^\s*\d{1,2}:\d{2}(:\d{2})?\s*$", na=False)

    dt = _robust_datetime_from_any(s)
    dt = dt.mask(time_only)

    out = dt.dt.strftime("%d/%m/%Y")
    out = out.where(~dt.isna(), np.nan)
    return out


def enforce_turno(series: pd.Series) -> pd.Series:
    """
    Normaliza TURNO a M/T/N:
    - MAÑANA/MANANA/M -> M
    - TARDE/T -> T
    - NOCHE/N -> N
    - DIA/D -> T (según tu regla)
    - valores tipo "L-M-V", "M-J-S", etc -> NaN
    """
    s = series.astype("object").astype(str).str.strip().str.upper()

    # horarios/semanas (no son turnos)
    bad = s.str.contains(r"^[LMDJVSD\-\/ ]{3,}$", regex=True, na=False)
    s = s.mask(bad)

    # normalizaciones comunes
    s = s.replace({
        "MAÑANA": "M", "MANANA": "M", "MANAÑA": "M",
        "TARDE": "T",
        "NOCHE": "N",
        "DIA": "T", "D": "T",
        "N": "N", "M": "M", "T": "T",
    })

    # si viene algo como "MT" o "TM" o "MN" etc -> indefinido, dejar vacío
    s = s.where(s.isin(["M", "T", "N"]), np.nan)
    return s


def normalize_codigo_value(v) -> str:
    """CODIGO como string numérico; recorta ceros a la izquierda; si >5 dígitos toma los últimos 5."""
    if v is None or (isinstance(v, float) and (np.isnan(v) or not np.isfinite(v))):
        return ""
    # si es número
    try:
        if isinstance(v, (int, np.integer)):
            return str(int(v))
        if isinstance(v, (float, np.floating)):
            if abs(float(v) - round(float(v))) < 1e-6:
                return str(int(round(float(v))))
            # raro: codigo con decimales -> extraer dígitos
    except Exception:
        pass

    s = str(v).strip()
    if s == "" or s.upper() in {"NAN","NONE"}:
        return ""
    # extraer dígitos
    digs = re.findall(r"\d+", s)
    if not digs:
        return s.strip().upper()  # lo reportaremos si no es numérico
    d = digs[0]
    d = d.lstrip("0") or "0"
    if len(d) > 5:
        d = d[-5:]
    return d


def repair_sf_shift(df: pd.DataFrame, report: list[str], context: str) -> pd.DataFrame:
    """
    En el caso bueno (DIC 2025), valores tipo SF113 (o similares) van en 'Unnamed: 1' (UBIGEO),
    NO en CODIGO.

    - Si CODIGO contiene un valor tipo SFxxxx -> se mueve a Unnamed: 1 si está vacío.
    - Si Unnamed: 1 ya tiene algo, igual se limpia CODIGO (para que CODIGO quede numérico).
    """
    if "CODIGO" not in df.columns or "Unnamed: 1" not in df.columns:
        return df

    df = df.copy()
    code = df["CODIGO"].astype("object")
    u1 = df["Unnamed: 1"].astype("object")

    code_s = code.astype(str).str.strip().str.upper()
    u1_s = u1.astype(str).str.strip()

    # ejemplos: SF113, SFI113, SFF113, etc.
    is_sf = code_s.str.match(r"^(SF\d{2,4}|S[A-Z]{0,2}F?I?\d{2,4})$", na=False)

    if not bool(is_sf.any()):
        return df

    empty_u1 = u1.isna() | (u1_s == "") | (u1_s.str.lower() == "nan")
    move = is_sf & empty_u1
    moved = int(move.sum())
    if moved:
        df.loc[move, "Unnamed: 1"] = code.loc[move]

    cleared = int(is_sf.sum())
    df.loc[is_sf, "CODIGO"] = np.nan

    report.append(f"[SHIFT] {context} | sf_en_codigo={cleared} moved_a_Unnamed1={moved}")
    return df



def normalize_codigo_column(df: pd.DataFrame, report: list[str], context: str) -> pd.DataFrame:
    """Normaliza CODIGO (00017024 -> 17024) y registra cambios."""
    if "CODIGO" not in df.columns:
        return df
    before = df["CODIGO"].astype("object")
    after = before.apply(normalize_codigo_value)
    # marcar solo donde cambió y after es numérico
    changed = (before.astype(str).str.strip() != after.astype(str).str.strip()) & (after != "")
    n = int(changed.sum())
    df = df.copy()
    df["CODIGO"] = after.replace("", np.nan)
    if n:
        report.append(f"[CODIGO] {context} | normalizados={n}")
    return df


def fill_codigo_with_mode(df: pd.DataFrame, report: list[str], context: str) -> pd.DataFrame:
    """
    Si CODIGO viene con valores inválidos (ej: vacíos, 'SF113', '00017024', etc),
    intenta rellenar SOLO los inválidos usando la moda de los códigos válidos de 5 dígitos
    dentro del mismo dataframe.
    """
    if "CODIGO" not in df.columns:
        return df

    df = df.copy()
    s = df["CODIGO"].astype("object").apply(normalize_codigo_value)
    s = s.replace("", np.nan)

    valid_mask = s.notna() & s.astype(str).str.match(r"^\d{5}$", na=False)
    if not bool(valid_mask.any()):
        df["CODIGO"] = s
        return df

    mode = s.loc[valid_mask].value_counts().idxmax()

    invalid_mask = s.isna() | (~s.astype(str).str.match(r"^\d{5}$", na=False))
    n = int(invalid_mask.sum())
    if n:
        s.loc[invalid_mask] = mode
        report.append(f"[CODIGO] {context} | rellenados_con_moda={n} (m={mode})")

    df["CODIGO"] = s
    return df


def _looks_like_cpt_str(s: str) -> bool:
    s = (s or "").strip().upper()
    if s == "" or s in {"NAN","NONE"}:
        return False
    s = s.replace(",", ".")
    s = re.sub(r"\s+", "", s)
    # letras+num(+dec)
    if re.fullmatch(r"[A-Z]+\d+(\.\d{1,3})?", s):
        return True
    # num(+dec)
    if re.fullmatch(r"\d{3,7}(\.\d{1,3})?", s):
        return True
    return False


def _is_weak_numeric_code(s: str) -> bool:
    """Códigos numéricos muy cortos (3 dígitos) suelen ser conteos (ej 160) cuando hay otros CPT."""
    s = (s or "").strip().upper()
    if re.fullmatch(r"\d{3}$", s):
        return True
    return False


def pack_cpt_block(df: pd.DataFrame, report: list[str], context: str, *args, **kwargs) -> pd.DataFrame:
    """
    Empaqueta el bloque CPT:
    - mueve el primer CPT real a 'PROCEDIMIENTO (CPT)'
    - desplaza opcionales
    - dedup dentro de la fila
    - NO destruye decimales/letras
    """
    cpt_cols = [c for c in df.columns if isinstance(c, str) and "PROCEDIMIENTO" in c.upper() and "CPT" in c.upper()]
    if not cpt_cols or "PROCEDIMIENTO   (CPT)" not in df.columns:
        return df

    df = df.copy()
    # normalizar valores de todas las columnas CPT
    for c in cpt_cols:
        df[c] = df[c].astype("object").apply(normalize_cpt_value).replace("", np.nan)

    packed_rows = 0
    changed_main = 0

    # iterar filas (vectorizar sería ideal, pero mantener simple y seguro)
    for i in range(len(df)):
        vals = []
        for c in cpt_cols:
            v = df.at[i, c]
            if pd.isna(v):
                continue
            vals.append(str(v).strip().upper())
        if not vals:
            continue

        # separar códigos y otros
        codes = [v for v in vals if _looks_like_cpt_str(v)]
        others = [v for v in vals if not _looks_like_cpt_str(v)]

        if not codes:
            continue

        # si hay algún código "fuerte", descartar los weak (3 dígitos) de la lista de códigos
        strong = [v for v in codes if not _is_weak_numeric_code(v)]
        if strong:
            weak = [v for v in codes if _is_weak_numeric_code(v)]
            codes = strong
            # los weak pasan a others para no perderlos
            others = weak + others

        # dedup preservando orden
        def dedup(seq):
            seen=set()
            out=[]
            for x in seq:
                if x in seen:
                    continue
                seen.add(x)
                out.append(x)
            return out
        codes = dedup(codes)
        others = dedup(others)

        # nuevo contenido para columnas CPT
        new_vals = (codes + others)[:len(cpt_cols)]
        # completar con NaN
        new_vals += [np.nan]*(len(cpt_cols)-len(new_vals))

        # detectar cambio en main CPT
        old_main = df.at[i, "PROCEDIMIENTO   (CPT)"]
        new_main = new_vals[cpt_cols.index("PROCEDIMIENTO   (CPT)")] if "PROCEDIMIENTO   (CPT)" in cpt_cols else new_vals[0]
        if (pd.isna(old_main) and pd.notna(new_main)) or (pd.notna(old_main) and pd.notna(new_main) and str(old_main).strip().upper() != str(new_main).strip().upper()):
            changed_main += 1

        # aplicar
        for c, v in zip(cpt_cols, new_vals):
            df.at[i, c] = v
        packed_rows += 1

    if packed_rows:
        report.append(f"[CPT] {context} | filas_revisadas={packed_rows} main_cambiado={changed_main}")
    return df


def split_by_categoria(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Separa registros por CATEGORIA DEL PERSONAL DE SALUD:
    - MED: 1-4
    - ENF: 5-7
    - UNK: otros o NaN
    """
    col = "CATEGORIA DEL PERSONAL DE SALUD"
    if col not in df.columns:
        empty = df.iloc[0:0].copy()
        return empty, empty, df.copy()

    cat = pd.to_numeric(df[col], errors="coerce")
    med = df[cat.isin([1,2,3,4])].copy()
    enf = df[cat.isin([5,6,7])].copy()
    unk = df[~cat.isin([1,2,3,4,5,6,7])].copy()
    return med, enf, unk


# ============================================================
# Detección de header "plantilla-compatible"
# ============================================================

def find_strict_header_row_openpyxl(ws, max_rows=60) -> Optional[int]:
    """
    Devuelve header_row (0-based) si existe una fila donde aparezcan
    ambos tokens: TURNO y CODIGO.
    """
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=0):
        vals = [_norm_txt(v) for v in row]
        sset = {v for v in vals if v}
        if "TURNO" in sset and "CODIGO" in sset:
            return r_idx
    return None


def detect_header_row_fallback_openpyxl(ws, max_rows=80) -> Optional[int]:
    """
    Fallback por score de tokens (para formatos no estándar).
    NOTA: si el archivo tiene una hoja buena con TURNO+CODIGO, NO se usa esto.
    """
    best_row, best_score = 0, -1
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=0):
        vals = [_norm_txt(v) for v in row]
        joined = " ".join(vals)
        score = sum(
            1 for t in ("TURNO","CODIGO","CATEGORIA","DNI","PROCEDIMIENTO","DIAGNOSTICO","FECHA","UBIGEO")
            if t in joined
        )
        if score > best_score:
            best_score, best_row = score, r_idx
    return None if best_score < 2 else best_row
def map_columns_to_template(df_cols: List[str], template_cols: List[str], fuzzy_threshold: float = 0.86) -> dict:
    """
    Mapea columnas del DF al nombre exacto de columna de la plantilla:
    - match exacto normalizado
    - fuzzy match (no Unnamed)
    """
    tpl_norm = {c: _norm_txt(c) for c in template_cols}
    mapping, used = {}, set()

    # exact
    for c in df_cols:
        nc = _norm_txt(c)
        for t, nt in tpl_norm.items():
            if nt and nt == nc and t not in used:
                mapping[c] = t
                used.add(t)
                break

    # fuzzy
    for c in df_cols:
        if c in mapping:
            continue
        nc = _norm_txt(c)
        if not nc or nc.startswith("UNNAMED"):
            continue

        best, best_score = None, 0.0
        for t in template_cols:
            if t in used:
                continue
            nt = tpl_norm[t]
            if not nt or nt.startswith("UNNAMED"):
                continue
            score = SequenceMatcher(None, nc, nt).ratio()
            if score > best_score:
                best_score, best = score, t

        if best is not None and best_score >= fuzzy_threshold:
            mapping[c] = best
            used.add(best)

    return mapping




def _fmt_num_code(x: float) -> str:
    """Formato estable para números tipo código: enteros sin .0; decimales a 2 dígitos."""
    if x is None or (isinstance(x, float) and (math.isnan(x) or not math.isfinite(x))):
        return ""
    if abs(x - round(x)) < 1e-6:
        return str(int(round(x)))
    # 2 decimales, preserva .01 / .10
    return f"{x:.2f}"


def normalize_cpt_value(v) -> str:
    """
    Normaliza un valor de CPT sin destruir:
    - conserva letras (C2062, D2392)
    - conserva decimales (99199.11, 99206.01)
    - corrige artefactos de float (.0, 99199.1100000001)
    - unifica separador decimal ',' -> '.'
    """
    if v is None:
        return ""
    try:
        import numpy as _np
        if isinstance(v, (_np.floating, float)):
            return _fmt_num_code(float(v))
        if isinstance(v, (_np.integer, int)):
            return str(int(v))
    except Exception:
        pass

    s = str(v).strip().upper()
    if s in {"", "NAN", "NONE"}:
        return ""
    s = s.replace(",", ".")
    s = re.sub(r"\s+", "", s)

    # quitar .0 exacto
    s = re.sub(r"\.0$", "", s)

    # Si es numérico puro con decimal:
    m = re.fullmatch(r"(\d+)\.(\d+)", s)
    if m:
        a, b = m.group(1), m.group(2)
        if len(b) == 1:
            return f"{a}.{b}0"
        if len(b) == 2:
            return s
        # demasiados decimales -> redondear a 2
        try:
            x = float(s)
            return _fmt_num_code(x)
        except Exception:
            return f"{a}.{b[:2]}"

    # Letra + num (+ decimal)
    m = re.fullmatch(r"([A-Z]+)(\d+)\.(\d+)", s)
    if m:
        pre, a, b = m.group(1), m.group(2), m.group(3)
        if len(b) == 1:
            return f"{pre}{a}.{b}0"
        if len(b) >= 2:
            return f"{pre}{a}.{b[:2]}"
    # Letra+num sin decimal
    m = re.fullmatch(r"([A-Z]+)(\d+)", s)
    if m:
        return f"{m.group(1)}{m.group(2)}"

    # num puro
    if re.fullmatch(r"\d+", s):
        return s

    # devolver tal cual (por si viniera un código raro); packing decidirá si usa o no
    return s


def clean_cpt(series: pd.Series) -> pd.Series:
    """
    Limpieza NO destructiva de CPT.
    Mantiene letras y decimales (como el caso bueno DIC 2025).
    """
    s = series.astype("object")
    out = s.apply(normalize_cpt_value)
    out = out.replace("", np.nan)
    return out


def clean_cie10(series: pd.Series) -> pd.Series:
    """
    Limpia CIE10:
    - K07,2 -> K07.2
    - R.10 -> R10
    - J00.X -> J00
    Mantiene formato general: Letra + 2 dígitos + opcional . + 1-4 alfanum.
    """
    s = series.astype("object").astype(str).str.strip().str.upper()
    s = s.replace({"NAN": ""})
    s = s.str.replace(",", ".", regex=False)
    s = s.str.replace(r"\s+", "", regex=True)
    # "R.10" -> "R10"
    s = s.str.replace(r"^([A-Z])\.(\d{2})", r"\1\2", regex=True)

    # extraer forma canonical
    m = s.str.extract(r"^([A-Z]\d{2})(?:\.?([A-Z0-9]{1,4}))?", expand=True)
    base = m[0]
    ext = m[1]

    out = base.copy()
    out = out.where(base.notna(), np.nan)

    # si ext es 'X' o termina en X, suele ser placeholder -> dejar solo base
    ext2 = ext.where(ext.notna(), "")
    ext2 = ext2.where(~ext2.str.fullmatch(r"X+"), "")

    out = out.where(ext2.eq(""), base + "." + ext2)
    return out.where(out.notna() & (out.astype(str).str.len() >= 3), np.nan)
# ============================================================
# Limpieza / reparación de columnas
# ============================================================

def drop_garbage_rows(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all").copy()

    # headers repetidos dentro del cuerpo
    if "TURNO" in df.columns:
        df = df[df["TURNO"].astype(str).str.upper().ne("TURNO")]

    # quitar filas completamente vacías en campos clave
    key_candidates = [c for c in [
        "Unnamed: 0", "TURNO", "CODIGO", "DNI DEL PACIENTE",
        "DIAGNOSTICO 1 (CIE 10)", "PROCEDIMIENTO   (CPT)"
    ] if c in df.columns]
    if key_candidates:
        df = df[~df[key_candidates].isna().all(axis=1)].copy()

    return df





def fill_unidad_origen_from_unnamed(df: pd.DataFrame, template_cols: List[str]) -> pd.DataFrame:
    """
    Si UNIDAD DE ORIGEN está vacía pero existe otra columna (a menudo Unnamed:xx)
    con valores tipo "EMERGENCIA", "CONSULTA EXTERNA", "CENTRO QUIRURGICO", etc.,
    la copia.
    """
    df = df.copy()
    target = "UNIDAD DE ORIGEN"
    if target not in template_cols:
        return df

    if target not in df.columns:
        df[target] = pd.Series([np.nan]*len(df), index=df.index, dtype="object")
    else:
        df[target] = df[target].astype("object")

    non_null_ratio = df[target].notna().mean() if len(df) else 0.0
    # Antes se retornaba si tenía >15% lleno, pero eso deja muchos huecos.
    # Solo retornamos si ya está prácticamente completo.
    if non_null_ratio > 0.97:
        return df

    keywords = ["EMERGENCIA","CONSULTA","CENTRO","QUIRURG","HOSPITAL","DPTO","SERVICIO","TRIAJE","HEMODIAL","ESTOMAT","UCI","URGEN"]
    best_col, best_score = None, 0.0

    # candidates: unnamed, no-template, o columnas típicamente mal corridas
    extra_candidates = {"ESPECIALIDAD DE MEDICO TRATANTE"}

    for c in df.columns:
        if c == target:
            continue
        if not (str(c).startswith("Unnamed") or c not in template_cols or c in extra_candidates):
            continue
        s = df[c]
        if s.isna().all():
            continue
        sample = s.dropna().astype(str).str.upper().head(600)
        if len(sample) < 10:
            continue
        # score por MAX keyword ratio (más robusto que promedio)
        ratios = [sample.str.contains(k, na=False).mean() for k in keywords]
        score = max(ratios) if ratios else 0.0
        if score > best_score:
            best_score, best_col = score, c

    if best_col is not None and best_score >= 0.06:
        miss = df[target].isna() | (df[target].astype(str).str.strip() == "")
        df.loc[miss, target] = df.loc[miss, best_col].astype("object").astype(str).str.strip()
        # limpiar 'nan'
        df[target] = df[target].replace({"nan": np.nan, "NAN": np.nan})

    return df


def apply_column_aliases(df: pd.DataFrame, report: list[str], context: str) -> pd.DataFrame:
    """
    Algunos excels no usan exactamente los mismos encabezados:
    - FECHA / FECHA DE ATENCION -> 'Unnamed: 0'
    - UBIGEO -> 'Unnamed: 1'
    - DNI PERSONAL / DNI TRATANTE -> 'Unnamed: 9'

    Esto reduce pérdidas de datos cuando el header no coincide con la plantilla.
    """
    df = df.copy()
    cols = list(df.columns)
    # --- Si el encabezado de la primera columna es una fecha (a veces ponen la FECHA en el header)
    #     renombrarla a Unnamed: 0 para que el resto del pipeline (ffill/fechas) funcione.
    if "Unnamed: 0" not in df.columns and len(df.columns) > 0:
        c0 = df.columns[0]
        if isinstance(c0, (datetime.datetime, datetime.date, pd.Timestamp)):
            df = df.rename(columns={c0: "Unnamed: 0"})
            report.append(f"[ALIAS] {context} | header_fecha -> Unnamed: 0")
    # --- Si existe una columna con nombre-fecha y Unnamed:0 no existe, intentar usar la primera fecha
    if "Unnamed: 0" not in df.columns:
        for c in df.columns:
            if isinstance(c, (datetime.datetime, datetime.date, pd.Timestamp)):
                df = df.rename(columns={c: "Unnamed: 0"})
                report.append(f"[ALIAS] {context} | header_fecha({c}) -> Unnamed: 0")
                break
    norm_map = {c: _norm_txt(c) for c in cols}

    def find_by_all_tokens(tokens: list[str]) -> Optional[str]:
        for c, n in norm_map.items():
            if all(t in n for t in tokens):
                return c
        return None

    # ---- FECHA -> Unnamed: 0
    if "Unnamed: 0" not in df.columns:
        c = find_by_all_tokens(["FECHA"])
        if c is not None:
            df = df.rename(columns={c: "Unnamed: 0"})
            report.append(f"[ALIAS] {context} | {c} -> Unnamed: 0")
    else:
        # si Unnamed:0 está muy vacío y existe FECHA, rellenar vacíos
        c = find_by_all_tokens(["FECHA"])
        if c is not None and c != "Unnamed: 0":
            missing = df["Unnamed: 0"].isna()
            if float(missing.mean()) > 0.25:
                df.loc[missing, "Unnamed: 0"] = df.loc[missing, c]
                report.append(f"[ALIAS] {context} | rellenado Unnamed:0 desde {c} (faltantes={int(missing.sum())})")

    # ---- UBIGEO -> Unnamed: 1
    if "Unnamed: 1" not in df.columns:
        c = find_by_all_tokens(["UBIGEO"])
        if c is not None:
            df = df.rename(columns={c: "Unnamed: 1"})
            report.append(f"[ALIAS] {context} | {c} -> Unnamed: 1")

    # ---- DNI PERSONAL -> Unnamed: 9
    if "Unnamed: 9" not in df.columns:
        c = find_by_all_tokens(["DNI", "PERSONAL"]) or find_by_all_tokens(["DNI", "TRATANTE"])
        if c is not None:
            df = df.rename(columns={c: "Unnamed: 9"})
            report.append(f"[ALIAS] {context} | {c} -> Unnamed: 9")

    # ---- UPSS (o similar) -> UNIDAD DE ORIGEN
    if "UNIDAD DE ORIGEN" not in df.columns:
        c = find_by_all_tokens(["UPSS"]) or find_by_all_tokens(["UNIDAD","ORIGEN"]) or find_by_all_tokens(["SERVICIO"]) 
        if c is not None:
            df = df.rename(columns={c: "UNIDAD DE ORIGEN"})
            report.append(f"[ALIAS] {context} | {c} -> UNIDAD DE ORIGEN")

    return df


def forward_fill_blocks(df: pd.DataFrame, cols) -> pd.DataFrame:
    """
    Muchos excels vienen con celdas fusionadas: el primer registro tiene FECHA/CODIGO/TURNO y
    el resto de filas del bloque vienen en blanco.

    - Soporta items tipo "COL" o tuplas/listas de alternativas: ("Unnamed: 0", "FECHA")
    - Trata blancos/espacios como NaN para que el ffill funcione.
    """
    df = df.copy()

    for item in cols:
        if isinstance(item, (tuple, list)):
            cand = [c for c in item if c in df.columns]
        else:
            cand = [item] if item in df.columns else []

        for c in cand:
            s = df[c].astype("object")
            # blancos -> NaN
            s = s.replace(r"^\s*$", np.nan, regex=True)
            s = s.replace({"nan": np.nan, "NAN": np.nan, "None": np.nan, "NONE": np.nan})
            df[c] = s.ffill()

    return df




def numeric_ratio(series: pd.Series, sample=400) -> float:
    s = series.dropna()
    if len(s) == 0:
        return 0.0
    if len(s) > sample:
        s = s.sample(sample, random_state=0)
    st = s.astype(str).str.strip()
    return st.str.match(r"^\d+(\.\d+)?$").mean()


def fix_procedure_pairs(df: pd.DataFrame, template_cols: List[str]) -> pd.DataFrame:
    """
    En varias plantillas, el CPT/código cae en la columna Unnamed contigua.
    Si detecta que el "código" está en Unnamed y el texto en PROCEDIMIENTO -> swap.
    """
    df = df.copy()
    for i, col in enumerate(template_cols[:-1]):
        nxt = template_cols[i + 1]
        if (
            isinstance(col, str) and col.startswith("PROCEDIMIENTO")
            and isinstance(nxt, str) and nxt.startswith("Unnamed")
            and col in df.columns and nxt in df.columns
        ):
            r1 = numeric_ratio(df[col])
            r2 = numeric_ratio(df[nxt])
            if r1 < 0.35 and r2 > 0.55:
                df[col], df[nxt] = df[nxt], df[col]
    return df


# ============================================================
# Bases de datos (corrección)
# ============================================================

def load_databases(pacientes_xlsx: str, personal_xlsx: str):
    pac = pd.read_excel(pacientes_xlsx)
    if "DNI" in pac.columns:
        pac["DNI"] = pac["DNI"].apply(normalize_dni)
    pac = pac.dropna(subset=["DNI"]).drop_duplicates(subset=["DNI"], keep="first")
    pac_lookup = pac.set_index("DNI").to_dict("index")

    per = pd.read_excel(personal_xlsx, header=2)
    per = per[per["Unnamed: 0"].astype(str).str.strip().ne("N°")].copy()
    per = per.rename(columns={
        "Unnamed: 1": "GRADO",
        "Unnamed: 2": "ESPECIALIDAD",
        "Unnamed: 3": "APELLIDOS_Y_NOMBRES",
        "Unnamed: 5": "DNI",
    })
    per["DNI"] = per["DNI"].apply(normalize_dni)
    per = per.dropna(subset=["DNI"]).drop_duplicates(subset=["DNI"], keep="first")
    per_lookup = per.set_index("DNI").to_dict("index")

    return pac_lookup, per_lookup


def enrich_with_dbs(df: pd.DataFrame, pac_lookup, per_lookup) -> pd.DataFrame:
    df = df.copy()

    # Paciente
    if "DNI DEL PACIENTE" in df.columns:
        df["DNI DEL PACIENTE"] = df["DNI DEL PACIENTE"].apply(normalize_dni)

    if "DNI DEL TITULAR" in df.columns:
        df["DNI DEL TITULAR"] = df["DNI DEL TITULAR"].apply(normalize_dni)

    if "Edad" in df.columns and "DNI DEL PACIENTE" in df.columns:
        df["Edad"] = df["Edad"].astype("object")
        miss = df["Edad"].isna() | (df["Edad"].astype(str).str.strip() == "")
        df.loc[miss, "Edad"] = df.loc[miss, "DNI DEL PACIENTE"].map(
            lambda d: pac_lookup.get(d, {}).get("edad") if pd.notna(d) else np.nan
        )

    if "Sexo" in df.columns and "DNI DEL PACIENTE" in df.columns:
        df["Sexo"] = df["Sexo"].astype("object")
        miss = df["Sexo"].isna() | (df["Sexo"].astype(str).str.strip() == "")
        df.loc[miss, "Sexo"] = df.loc[miss, "DNI DEL PACIENTE"].map(
            lambda d: pac_lookup.get(d, {}).get("Sexo") if pd.notna(d) else np.nan
        )

    # Completar campos adicionales del paciente si existen en la plantilla
    if "Grupo" in df.columns and "DNI DEL PACIENTE" in df.columns:
        df["Grupo"] = df["Grupo"].astype("object")
        miss = df["Grupo"].isna() | (df["Grupo"].astype(str).str.strip() == "")
        df.loc[miss, "Grupo"] = df.loc[miss, "DNI DEL PACIENTE"].map(
            lambda d: pac_lookup.get(d, {}).get("Categoria") if pd.notna(d) else np.nan
        )

    if "Situación" in df.columns and "DNI DEL PACIENTE" in df.columns:
        df["Situación"] = df["Situación"].astype("object")
        miss = df["Situación"].isna() | (df["Situación"].astype(str).str.strip() == "")
        df.loc[miss, "Situación"] = df.loc[miss, "DNI DEL PACIENTE"].map(
            lambda d: pac_lookup.get(d, {}).get("estado") if pd.notna(d) else np.nan
        )

    if "Condición" in df.columns and "DNI DEL PACIENTE" in df.columns:
        df["Condición"] = df["Condición"].astype("object")
        miss = df["Condición"].isna() | (df["Condición"].astype(str).str.strip() == "")
        df.loc[miss, "Condición"] = df.loc[miss, "DNI DEL PACIENTE"].map(
            lambda d: pac_lookup.get(d, {}).get("Esta_Ate") if pd.notna(d) else np.nan
        )

    # Personal: si existe DNI del personal en la columna Unnamed:9, corrige nombre/especialidad
    if "Unnamed: 9" in df.columns:
        df["Unnamed: 9"] = df["Unnamed: 9"].apply(normalize_dni)

        name_col = "NOMBRE DEL MEDICO / PERSONAL DE SALUD TRATANTE"
        esp_col = "ESPECIALIDAD DE MEDICO TRATANTE"

        if name_col in df.columns:
            df[name_col] = df.apply(
                lambda r: per_lookup.get(r["Unnamed: 9"], {}).get("APELLIDOS_Y_NOMBRES")
                if pd.notna(r.get("Unnamed: 9")) and r.get("Unnamed: 9") in per_lookup
                else r.get(name_col),
                axis=1
            )

        if esp_col in df.columns:
            df[esp_col] = df.apply(
                lambda r: per_lookup.get(r["Unnamed: 9"], {}).get("ESPECIALIDAD")
                if pd.notna(r.get("Unnamed: 9")) and r.get("Unnamed: 9") in per_lookup
                else r.get(esp_col),
                axis=1
            )

    return df


# ============================================================
# Duplicados entre hojas del mismo archivo (rápido)
# ============================================================


def _decat_df(df: pd.DataFrame) -> pd.DataFrame:
    """Convert any categorical columns to plain object to avoid sort errors."""
    for c in df.columns:
        try:
            if pd.api.types.is_categorical_dtype(df[c]):
                df[c] = df[c].astype("object")
        except Exception:
            pass
    return df
def quick_fingerprint(df: pd.DataFrame, key_cols: List[str], sample_each=120) -> str:
    """
    Huella rápida: hash de tuplas (primeras/últimas filas) en columnas clave.
    Evita comparar DF completos (caro).
    """
    if not key_cols:
        # fallback: shape+columns
        base = f"{df.shape}|{','.join(map(str, df.columns))}"
        return hashlib.md5(base.encode("utf-8")).hexdigest()

    sub = df[key_cols].copy()
    for c in key_cols:
        sub[c] = sub[c].astype(str).replace("nan", "").str.strip()

    head = sub.head(sample_each)
    tail = sub.tail(sample_each)
    mix = pd.concat([head, tail], ignore_index=True)

    joined = "\n".join("|".join(map(str, row)) for row in mix.itertuples(index=False, name=None))
    return hashlib.md5(joined.encode("utf-8")).hexdigest()


# ============================================================
# Escritura en plantilla (sin romper formato)
# ============================================================

def _find_last_nonempty_row(ws, start_row=3) -> int:
    max_r = ws.max_row
    max_c = ws.max_column
    for r in range(max_r, start_row - 1, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, max_c + 1)):
            return r
    return start_row - 1


def _clear_data_area(ws, start_row=3):
    last = _find_last_nonempty_row(ws, start_row)
    if last < start_row:
        return
    for r in range(start_row, last + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None


def _write_df(ws, df: pd.DataFrame, start_row=3):
    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, val in enumerate(row, start=1):
            if isinstance(val, float) and np.isnan(val):
                val = None
            ws.cell(i, j).value = val


# ============================================================
# Procesamiento por grupo (MED y ENF por separado)
# ============================================================

def _drop_obvious_empty_records(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtro NO agresivo: elimina solo filas totalmente vacías/obvias.
    Importante: considera CPT en columnas opcionales (PROC. ENF suele ponerlo ahí).
    """
    df = df.copy()

    # columnas clínicas dinámicas
    cpt_cols = [c for c in df.columns if isinstance(c, str) and ("PROCEDIMIENTO" in c.upper()) and ("CPT" in c.upper())]
    diag_cols = [c for c in df.columns if isinstance(c, str) and ("DIAGNOSTICO" in c.upper())]
    id_cols  = [c for c in ["CODIGO","DNI DEL PACIENTE","DNI DEL TITULAR"] if c in df.columns]

    # keys mínimas para descartar filas vacías
    keys = [c for c in ["Unnamed: 0","TURNO","CODIGO"] if c in df.columns] + id_cols + cpt_cols[:3] + diag_cols[:2]
    keys = [c for c in keys if c in df.columns]
    if keys:
        all_empty = df[keys].isna().all(axis=1)
        df = df.loc[~all_empty].copy()

    # descartar "basura" cuando no hay turno ni ids ni clínicos
    if "TURNO" in df.columns:
        clin_cols = (cpt_cols + diag_cols)
        if id_cols and clin_cols:
            bad = df["TURNO"].isna() & df[id_cols].isna().all(axis=1) & df[clin_cols].isna().all(axis=1)
            df = df.loc[~bad].copy()

    return df






def _process_files(
    paths: List[str],
    tpl_med_cols: List[str],
    tpl_enf_cols: List[str],
    pac_lookup,
    per_lookup,
    report: List[str],
    group_name: str,
    sort_output: bool,
    route_by_category: bool,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Procesa un conjunto de archivos (sean MED o ENF seleccionados por el usuario),
    pero rutea filas al output correcto usando CATEGORIA DEL PERSONAL DE SALUD.
    Esto evita mezclas y corrige cuando un archivo viene "mixto".
    """
    med_parts = []
    enf_parts = []

    super_cols = list(dict.fromkeys(tpl_med_cols + tpl_enf_cols))  # unión, preserva orden

    for path in paths:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
        except Exception as e:
            report.append(f"[{group_name}] [ERR] {os.path.basename(path)} | No se pudo abrir: {type(e).__name__}")
            continue

        strict_sheets: List[Tuple[str, int]] = []
        fallback_sheets: List[Tuple[str, int]] = []

        for sh in wb.sheetnames:
            ws = wb[sh]
            hr = find_strict_header_row_openpyxl(ws)
            if hr is not None:
                strict_sheets.append((sh, hr))
            else:
                fr = detect_header_row_fallback_openpyxl(ws)
                if fr is not None:
                    fallback_sheets.append((sh, fr))

        chosen = strict_sheets if strict_sheets else fallback_sheets

        if strict_sheets:
            for sh, _fr in fallback_sheets:
                report.append(f"[{group_name}] [SKIP] {os.path.basename(path)} :: {sh} | Hoja resumen/no-plantilla (hay otra hoja con TURNO+CODIGO en el mismo archivo)")

        seen_fp = set()

        for sh, hr in chosen:
            try:
                df = pd.read_excel(path, sheet_name=sh, header=hr)
            except Exception as e:
                report.append(f"[{group_name}] [ERR] {os.path.basename(path)} :: {sh} | ReadError {type(e).__name__}")
                continue

            in_rows = len(df)

            # mapear a super esquema (reduce vacíos)
            mapping = map_columns_to_template(list(df.columns), super_cols)
            if mapping:
                df = df.rename(columns=mapping)

            context = f"{os.path.basename(path)} :: {sh}"
            df = apply_column_aliases(df, report, context)
            df = drop_garbage_rows(df)

            # fechas/turno
            if "Unnamed: 0" in df.columns:
                df["Unnamed: 0"] = standardize_date_contextual(df["Unnamed: 0"], report, f"{os.path.basename(path)} :: {sh}")
            if "TURNO" in df.columns:
                df["TURNO"] = enforce_turno(df["TURNO"])

            # unidad y forward-fill por celdas fusionadas
            df = fill_unidad_origen_from_unnamed(df, super_cols)
            df = forward_fill_blocks(df, [
                ("Unnamed: 0", "FECHA"),
                "TURNO",
                "CODIGO",
                "CATEGORIA DEL PERSONAL DE SALUD",
                "UNIDAD DE ORIGEN",
                ("Unnamed: 9",),
            ])

            # Reparar desplazamientos tipo SF113 (columna errónea) y normalizar CODIGO
            df = repair_sf_shift(df, report, f"{os.path.basename(path)} :: {sh}")
            df = normalize_codigo_column(df, report, f"{os.path.basename(path)} :: {sh}")

            # Limpiar CPT (no destructivo) para TODAS las columnas CPT existentes
            for _c in [c for c in df.columns if isinstance(c, str) and "PROCEDIMIENTO" in c.upper() and "CPT" in c.upper()]:
                df[_c] = clean_cpt(df[_c])

            # limpiar códigos
            if "DIAGNOSTICO 1 (CIE 10)" in df.columns:
                df["DIAGNOSTICO 1 (CIE 10)"] = clean_cie10(df["DIAGNOSTICO 1 (CIE 10)"])

            # ensure columns for db fill (si no vienen en el excel, igual se pueden completar con las BD)
            for _c in ("Edad","Sexo","Grupo","Situación","Condición",
                       "NOMBRE DEL MEDICO / PERSONAL DE SALUD TRATANTE",
                       "ESPECIALIDAD DE MEDICO TRATANTE"):
                if _c in super_cols and _c not in df.columns:
                    df[_c] = np.nan

            df = enrich_with_dbs(df, pac_lookup, per_lookup)

            # fingerprint para evitar duplicados dentro de un mismo archivo
            fp_cols = [c for c in ["Unnamed: 0","TURNO","CODIGO","DNI DEL PACIENTE","PROCEDIMIENTO   (CPT)","DIAGNOSTICO 1 (CIE 10)"] if c in df.columns]
            fp = quick_fingerprint(df.dropna(how="all"), fp_cols)
            if fp in seen_fp:
                report.append(f"[{group_name}] [SKIP] {os.path.basename(path)} :: {sh} | Duplicado (misma data en otra hoja) | in={in_rows} out={len(df)}")
                continue
            seen_fp.add(fp)

            # Distribución de categorías (solo para QA)
            cat_col = "CATEGORIA DEL PERSONAL DE SALUD"
            cat = pd.to_numeric(df[cat_col], errors="coerce") if cat_col in df.columns else pd.Series([np.nan]*len(df), index=df.index)
            cnt_med = int(cat.isin([1,2,3,4]).sum())
            cnt_enf = int(cat.isin([5,6,7]).sum())
            cnt_unk = int((~cat.isin([1,2,3,4,5,6,7]) | cat.isna()).sum())

            if route_by_category:
                # ruteo real por categoria (modo anterior)
                med, enf, unk = split_by_categoria(df)
                if len(unk):
                    if group_name == "MED_FILES":
                        med = pd.concat([med, unk], ignore_index=True)
                    else:
                        enf = pd.concat([enf, unk], ignore_index=True)
                    report.append(f"[{group_name}] [WARN] {os.path.basename(path)} :: {sh} | unk_categoria={len(unk)} (ruteado por grupo)")

                if len(med):
                    med2 = fix_procedure_pairs(med, tpl_med_cols).reindex(columns=tpl_med_cols)
                    med2 = _drop_obvious_empty_records(med2)
                    med2 = pack_cpt_block(med2, report, f"{os.path.basename(path)} :: {sh} :: PROC.MED")
                    med_parts.append(med2)

                if len(enf):
                    enf2 = fix_procedure_pairs(enf, tpl_enf_cols).reindex(columns=tpl_enf_cols)
                    enf2 = _drop_obvious_empty_records(enf2)
                    enf2 = pack_cpt_block(enf2, report, f"{os.path.basename(path)} :: {sh} :: PROC.ENF")
                    enf_parts.append(enf2)

                report.append(f"[{group_name}] [OK] {os.path.basename(path)} :: {sh} | in={in_rows} med={len(med)} enf={len(enf)} | cats_med={cnt_med} cats_enf={cnt_enf} cats_unk={cnt_unk}")

            else:
                # Modo solicitado: RESPETAR el grupo de carga del usuario
                target_cols = tpl_med_cols if group_name == "MED_FILES" else tpl_enf_cols
                out = fix_procedure_pairs(df, target_cols).reindex(columns=target_cols)
                out = _drop_obvious_empty_records(out)
                out = pack_cpt_block(out, report, f"{os.path.basename(path)} :: {sh} :: {('PROC.MED' if group_name=='MED_FILES' else 'PROC.ENF')}")

                # Warn si el archivo parece del otro grupo
                if group_name == "MED_FILES" and cnt_enf > max(30, int(0.25*len(df))):
                    report.append(f"[{group_name}] [WARN] {os.path.basename(path)} :: {sh} | Muchas filas con categoria ENF (5-7) dentro de archivos MED: {cnt_enf}/{len(df)}")
                if group_name == "ENF_FILES" and cnt_med > max(30, int(0.25*len(df))):
                    report.append(f"[{group_name}] [WARN] {os.path.basename(path)} :: {sh} | Muchas filas con categoria MED (1-4) dentro de archivos ENF: {cnt_med}/{len(df)}")

                if group_name == "MED_FILES":
                    med_parts.append(out)
                else:
                    enf_parts.append(out)

                report.append(f"[{group_name}] [OK] {os.path.basename(path)} :: {sh} | in={in_rows} out={len(out)} | cats_med={cnt_med} cats_enf={cnt_enf} cats_unk={cnt_unk}")

        wb.close()

    med_df = pd.concat(med_parts, ignore_index=True) if med_parts else pd.DataFrame(columns=tpl_med_cols)
    enf_df = pd.concat(enf_parts, ignore_index=True) if enf_parts else pd.DataFrame(columns=tpl_enf_cols)

    # dedup cross-files
    for df, cols in ((med_df, tpl_med_cols), (enf_df, tpl_enf_cols)):
        pass

    # ordenar (seguro) si se desea
    if sort_output:
        def _sort(df: pd.DataFrame) -> pd.DataFrame:
            if len(df) == 0 or "Unnamed: 0" not in df.columns:
                return df
            dtp = _robust_datetime_from_any(df["Unnamed: 0"])
            df = df.copy()
            df["_d"] = dtp
            df["_t"] = df["TURNO"].map({"M": 0, "T": 1, "N": 2}) if "TURNO" in df.columns else np.nan
            df["_c"] = pd.to_numeric(df["CODIGO"], errors="coerce") if "CODIGO" in df.columns else np.nan
            df = _decat_df(df)
            df = df.sort_values(by=["_d","_t","_c"], na_position="last", kind="mergesort")
            return df.drop(columns=["_d","_t","_c"])
        med_df = _sort(med_df)
        enf_df = _sort(enf_df)

    return med_df, enf_df
# ============================================================
# API principal
# ============================================================

def consolidate(
    plantilla_xlsx: str,
    pacientes_xlsx: str,
    personal_xlsx: str,
    proc_med_files: List[str],
    proc_enf_files: List[str],
    out_xlsx: str,
    include_audit_sheet: bool = False,
    write_report_txt: bool = True,
    sort_output: bool = True,
    route_by_category: bool = False,
):
    """
    - PROC. MED. se llena SOLO con proc_med_files
    - PROC. ENF se llena SOLO con proc_enf_files
    - NO filtra por fecha/año (usa TODO lo que venga)
    - Evita hojas "resumen" que rompen el armado (si existe una hoja compatible con plantilla, usa solo esa)
    - Respeta formato de Plantilla.xlsx (se escribe en el mismo libro)
    - Reporte .txt opcional al lado del output
    """
    tpl_med_cols = pd.read_excel(plantilla_xlsx, sheet_name="PROC. MED.", header=1, nrows=0).columns.tolist()
    tpl_enf_cols = pd.read_excel(plantilla_xlsx, sheet_name="PROC. ENF", header=1, nrows=0).columns.tolist()

    pac_lookup, per_lookup = load_databases(pacientes_xlsx, personal_xlsx)

    report: List[str] = []
    report.append("=== CONSOLIDADOR HOSPITAL - REPORTE (v4.8.3) ===")

    med_a, enf_a = _process_files(proc_med_files or [], tpl_med_cols, tpl_enf_cols, pac_lookup, per_lookup, report, "MED_FILES", sort_output, route_by_category)
    med_b, enf_b = _process_files(proc_enf_files or [], tpl_med_cols, tpl_enf_cols, pac_lookup, per_lookup, report, "ENF_FILES", sort_output, route_by_category)

    # combinar (ya ruteado por categoria)
    med_df = pd.concat([med_a, med_b], ignore_index=True) if len(med_b) else med_a
    enf_df = pd.concat([enf_a, enf_b], ignore_index=True) if len(enf_b) else enf_a

    report.append("")
    def _date_outlier_summary(df: pd.DataFrame, label: str):
        if "Unnamed: 0" not in df.columns or len(df)==0:
            return
        dt = pd.to_datetime(df["Unnamed: 0"], errors="coerce", dayfirst=True)
        dt = dt.dropna()
        if len(dt) < 50:
            return
        med = dt.median()
        out = dt[dt > med + pd.Timedelta(days=120)]
        if len(out):
            vc = out.dt.to_period("M").astype(str).value_counts().head(6)
            report.append(f"[OUTLIERS] {label} | future>120d={len(out)} | top_months=" + ", ".join([f"{k}:{int(v)}" for k,v in vc.items()]))

    _date_outlier_summary(med_df, "PROC.MED")
    _date_outlier_summary(enf_df, "PROC.ENF")
    report.append(f"PROC. MED. filas: {len(med_df)}")
    report.append(f"PROC. ENF. filas: {len(enf_df)}")

    
    def _final_cleanup(df: pd.DataFrame, label: str) -> pd.DataFrame:
        df = df.copy()
        context = f"FINAL {label}"

        # aliases (por si entraron columnas FECHA/UBIGEO/etc)
        df = apply_column_aliases(df, report, context)

        # fecha
        if "Unnamed: 0" in df.columns:
            df["Unnamed: 0"] = standardize_date_contextual(df["Unnamed: 0"], report, context)

        # turno
        if "TURNO" in df.columns:
            df["TURNO"] = enforce_turno(df["TURNO"])

        # forward fill (celdas fusionadas)
        df = forward_fill_blocks(df, [
            ("Unnamed: 0", "FECHA"),
            "TURNO",
            "CODIGO",
            "CATEGORIA DEL PERSONAL DE SALUD",
            "UNIDAD DE ORIGEN",
            ("Unnamed: 9",),
        ])

        # arreglos CODIGO/UBIGEO (SFxxx)
        df = repair_sf_shift(df, report, context)
        df = normalize_codigo_column(df, report, context)
        df = fill_codigo_with_mode(df, report, context)

        # CPT: normaliza + pack para evitar que quede vacío el principal
        cpt_cols = [c for c in df.columns if isinstance(c, str) and ("CPT" in c.upper()) and ("PROCEDIMIENTO" in c.upper())]
        if cpt_cols:
            for c in cpt_cols:
                df[c] = df[c].astype("object").apply(normalize_cpt_value).replace("", np.nan)
            df = pack_cpt_block(df, cpt_cols, report, context)

        # descarta filas que no parecen registros (sin DNI paciente + sin diagnóstico + sin CPT)
        key_cols = [c for c in ["DNI DEL PACIENTE", "DIAGNOSTICO 1 (CIE 10)", "PROCEDIMIENTO   (CPT)"] if c in df.columns]
        if key_cols:
            empty_mask = df[key_cols].isna().all(axis=1)
            if bool(empty_mask.any()):
                n = int(empty_mask.sum())
                df = df.loc[~empty_mask].copy()
                report.append(f"[DROP] {context} | filas_sin_claves={n}")

        return df

    med_df = _final_cleanup(med_df, "PROC. MED.")
    enf_df = _final_cleanup(enf_df, "PROC. ENF")

# Escribir dentro de la plantilla para conservar TODO el formato
    wb = openpyxl.load_workbook(plantilla_xlsx, keep_links=False)
    ws_med = wb["PROC. MED."]
    ws_enf = wb["PROC. ENF"]

    _clear_data_area(ws_med, start_row=3)
    _clear_data_area(ws_enf, start_row=3)

    _write_df(ws_med, med_df, start_row=3)
    _write_df(ws_enf, enf_df, start_row=3)

    # Hoja AUDITORIA (opcional)
    if include_audit_sheet:
        if "AUDITORIA" in wb.sheetnames:
            del wb["AUDITORIA"]
        ws_a = wb.create_sheet("AUDITORIA")
        ws_a.append(["nota"])
        ws_a.append(["AUDITORIA está pensada para logs técnicos; usa el REPORTE.txt para ver qué hojas entraron o se saltaron."])
    else:
        if "AUDITORIA" in wb.sheetnames:
            del wb["AUDITORIA"]

    wb.save(out_xlsx)

    if write_report_txt:
        base, _ = os.path.splitext(out_xlsx)
        report_path = base + "_REPORTE.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(map(str, report)))

